"""Build the 12-week MMA Training Tracker Excel workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# ── Colors ──
BG_DARK = "0A0A0F"
BG_CARD = "111114"
BG_RAISED = "1A1A20"
BG_SURFACE = "1C1C24"
BORDER_CLR = "2A2A36"
TEXT = "D4D4D8"
TEXT_BRIGHT = "EDEDF0"
TEXT_DIM = "71717A"
ACCENT = "F97316"
GREEN = "22C55E"
YELLOW = "EAB308"
RED = "EF4444"
BLUE = "3B82F6"
WHITE = "FFFFFF"
BLACK = "000000"

# Fills
fill_bg = PatternFill("solid", fgColor=BG_DARK)
fill_card = PatternFill("solid", fgColor=BG_CARD)
fill_raised = PatternFill("solid", fgColor=BG_RAISED)
fill_surface = PatternFill("solid", fgColor=BG_SURFACE)
fill_accent = PatternFill("solid", fgColor=ACCENT)
fill_green = PatternFill("solid", fgColor="1A3A2A")
fill_yellow = PatternFill("solid", fgColor="3A3520")
fill_red = PatternFill("solid", fgColor="3A1A1A")
fill_header = PatternFill("solid", fgColor="18181F")
fill_input = PatternFill("solid", fgColor="141420")

# Fonts
font_title = Font(name="Calibri", size=18, bold=True, color=TEXT_BRIGHT)
font_subtitle = Font(name="Calibri", size=12, bold=True, color=ACCENT)
font_header = Font(name="Calibri", size=10, bold=True, color=TEXT_BRIGHT)
font_body = Font(name="Calibri", size=10, color=TEXT)
font_dim = Font(name="Calibri", size=9, color=TEXT_DIM)
font_input = Font(name="Calibri", size=10, color=WHITE)
font_accent = Font(name="Calibri", size=10, bold=True, color=ACCENT)
font_accent_sm = Font(name="Calibri", size=9, bold=True, color=ACCENT)
font_check = Font(name="Calibri", size=14, color=GREEN)
font_week_label = Font(name="Calibri", size=11, bold=True, color=TEXT_BRIGHT)

# Alignment
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Border
thin_border = Border(
    left=Side(style="thin", color=BORDER_CLR),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)


def style_cell(cell, font=font_body, fill=fill_bg, alignment=center, border=thin_border):
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def make_title_row(ws, row, col, text, merge_end_col):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    cell = ws.cell(row=row, column=col, value=text)
    style_cell(cell, font=font_title, fill=fill_card)
    for c in range(col, merge_end_col + 1):
        style_cell(ws.cell(row=row, column=c), font=font_title, fill=fill_card)


def make_subtitle_row(ws, row, col, text, merge_end_col):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    cell = ws.cell(row=row, column=col, value=text)
    style_cell(cell, font=font_subtitle, fill=fill_bg)
    for c in range(col, merge_end_col + 1):
        style_cell(ws.cell(row=row, column=c), font=font_subtitle, fill=fill_bg)


def fill_row_bg(ws, row, start_col, end_col, fill=fill_bg):
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = thin_border


# ═══════════════════════════════════════════════════
# Session schedule data (AM/PM for each day, Mon-Sat)
# ═══════════════════════════════════════════════════
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

SESSIONS = {
    # Week type -> [(AM, PM), ...] for Mon-Sat
    "normal": [
        ("Strength — Upper Push/Pull", "MMA Striking"),
        ("Cardiac Output Cardio", "MMA Grappling"),
        ("Strength — Lower + Posterior", "Light Technique / Open Mat"),
        ("HIIT / Fight Circuit", "MMA Sparring"),
        ("Strength — Full Body Power", "MMA Drilling"),
        ("Fight Conditioning", "Open Mat / Light Technical"),
    ],
    "deload": [
        ("Strength — Upper (Light)", "MMA Striking — Light Technical"),
        ("Cardiac Output (Easy)", "MMA Grappling — Flow Rolling"),
        ("Strength — Lower (Light)", "Mobility / Yoga / Recovery"),
        ("Light Conditioning", "MMA Technical Drilling"),
        ("Strength — Full Body (Light)", "MMA Drilling — Technique"),
        ("Active Recovery — Easy Cardio", "Open Mat / Light Technical"),
    ],
}

INTENSITY = {
    "normal_w1": ["Y", "G", "R", "R", "Y", "Y",   "R", "Y", "G", "R", "Y", "G"],
    "normal_w2": ["Y", "G", "R", "R", "Y", "Y",   "R", "Y", "G", "R", "Y", "G"],
    "normal_w3": ["R", "G", "R", "R", "R", "R",   "R", "R", "G", "R", "Y", "G"],
    "deload":    ["G", "G", "G", "G", "G", "G",   "Y", "G", "G", "Y", "G", "G"],
}

INTENSITY_FILL = {"G": fill_green, "Y": fill_yellow, "R": fill_red}
INTENSITY_LABEL = {"G": "Easy", "Y": "Mod", "R": "Hard"}
INTENSITY_FONT = {
    "G": Font(name="Calibri", size=8, color=GREEN),
    "Y": Font(name="Calibri", size=8, color=YELLOW),
    "R": Font(name="Calibri", size=8, color=RED),
}


def build_daily_tracker(wb):
    """Sheet 1: 12-week daily session tracker with checkboxes."""
    ws = wb.active
    ws.title = "Daily Tracker"
    ws.sheet_properties.tabColor = ACCENT

    # Set dark background for entire visible area
    for r in range(1, 200):
        for c in range(1, 20):
            style_cell(ws.cell(row=r, column=c), fill=fill_bg)

    set_col_widths(ws, [2, 14, 28, 6, 6, 28, 6, 6, 10, 2])
    # Cols: spacer | Day | AM Session | Done | RPE | PM Session | Done | RPE | Notes | spacer

    row = 2
    make_title_row(ws, row, 2, "MMA TRAINING — 12-WEEK DAILY TRACKER", 9)
    row += 1
    make_subtitle_row(ws, row, 2, "Mark sessions complete (X) and log your RPE (1-10)", 9)
    row += 2

    cycle_names = ["CYCLE 1", "CYCLE 2", "CYCLE 3"]
    week_themes = ["Base Volume", "Progressive Overload", "Peak Volume", "Deload"]

    for cycle_idx in range(3):
        # Cycle header
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=2, value=f"── {cycle_names[cycle_idx]} (Weeks {cycle_idx*4+1}–{cycle_idx*4+4}) ──")
        style_cell(cell, font=Font(name="Calibri", size=13, bold=True, color=ACCENT), fill=fill_card)
        for c in range(2, 10):
            ws.cell(row=row, column=c).fill = fill_card
            ws.cell(row=row, column=c).border = thin_border
        row += 2

        for week_in_cycle in range(4):
            week_num = cycle_idx * 4 + week_in_cycle + 1
            is_deload = (week_in_cycle == 3)
            week_type = "deload" if is_deload else "normal"

            if week_in_cycle == 0:
                intensity_key = "normal_w1"
            elif week_in_cycle == 1:
                intensity_key = "normal_w2"
            elif week_in_cycle == 2:
                intensity_key = "normal_w3"
            else:
                intensity_key = "deload"

            # Week label
            theme = week_themes[week_in_cycle]
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
            label = f"WEEK {week_num} — {theme}"
            if is_deload:
                label += " (Reduced Volume)"
            cell = ws.cell(row=row, column=2, value=label)
            style_cell(cell, font=font_week_label, fill=fill_raised)
            for c in range(2, 10):
                ws.cell(row=row, column=c).fill = fill_raised
                ws.cell(row=row, column=c).border = thin_border
            row += 1

            # Column headers
            headers = ["Day", "AM Session", "\u2713", "RPE", "PM Session", "\u2713", "RPE", "Notes"]
            for i, h in enumerate(headers):
                cell = ws.cell(row=row, column=2 + i, value=h)
                style_cell(cell, font=font_header, fill=fill_header)
            row += 1

            sessions = SESSIONS[week_type]
            intensities = INTENSITY[intensity_key]

            for day_idx, day in enumerate(DAYS):
                am_session, pm_session = sessions[day_idx]
                am_int = intensities[day_idx]
                pm_int = intensities[day_idx + 6]

                am_label = f"{am_session}"
                pm_label = f"{pm_session}"

                # Day
                cell = ws.cell(row=row, column=2, value=day)
                style_cell(cell, font=font_accent, fill=fill_card, alignment=left_wrap)

                # AM Session
                cell = ws.cell(row=row, column=3, value=am_label)
                style_cell(cell, font=font_body, fill=INTENSITY_FILL[am_int], alignment=left_wrap)

                # AM Done (input)
                cell = ws.cell(row=row, column=4)
                style_cell(cell, font=font_check, fill=fill_input)

                # AM RPE (input)
                cell = ws.cell(row=row, column=5)
                style_cell(cell, font=font_input, fill=fill_input)

                # PM Session
                cell = ws.cell(row=row, column=6, value=pm_label)
                style_cell(cell, font=font_body, fill=INTENSITY_FILL[pm_int], alignment=left_wrap)

                # PM Done (input)
                cell = ws.cell(row=row, column=7)
                style_cell(cell, font=font_check, fill=fill_input)

                # PM RPE (input)
                cell = ws.cell(row=row, column=8)
                style_cell(cell, font=font_input, fill=fill_input)

                # Notes (input)
                cell = ws.cell(row=row, column=9)
                style_cell(cell, font=font_input, fill=fill_input, alignment=left_wrap)

                row += 1

            # Sunday row
            cell = ws.cell(row=row, column=2, value="Sunday")
            style_cell(cell, font=font_dim, fill=fill_card, alignment=left_wrap)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
            cell = ws.cell(row=row, column=3, value="REST DAY")
            style_cell(cell, font=font_dim, fill=fill_card)
            for c in range(3, 10):
                ws.cell(row=row, column=c).fill = fill_card
                ws.cell(row=row, column=c).border = thin_border
            row += 1

            # Week completion summary formula
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=2, value="Sessions Completed:")
            style_cell(cell, font=font_accent_sm, fill=fill_bg)
            # Count non-empty cells in done columns
            done_start = row - 7  # 6 days + 1 sunday row back
            done_end = row - 2
            formula = f'=COUNTA(D{done_start}:D{done_end})+COUNTA(G{done_start}:G{done_end})&" / 12"'
            cell = ws.cell(row=row, column=6, value=formula)
            style_cell(cell, font=font_accent, fill=fill_bg)
            row += 2

        row += 1  # Extra space between cycles

    # Legend at bottom
    make_subtitle_row(ws, row, 2, "LEGEND", 9)
    row += 1
    legends = [
        ("Green cells = Easy / Recovery", fill_green, GREEN),
        ("Yellow cells = Moderate", fill_yellow, YELLOW),
        ("Red cells = Hard / High demand", fill_red, RED),
        ("Dark cells = Your input (type X for done, 1-10 for RPE)", fill_input, TEXT_DIM),
    ]
    for text, fill, color in legends:
        cell = ws.cell(row=row, column=2, value=text)
        style_cell(cell, font=Font(name="Calibri", size=9, color=color), fill=fill, alignment=left_wrap)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        for c in range(2, 10):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    ws.freeze_panes = "B6"


def build_strength_log(wb):
    """Sheet 2: Strength progression log for main lifts across 12 weeks."""
    ws = wb.create_sheet("Strength Log")
    ws.sheet_properties.tabColor = BLUE

    for r in range(1, 80):
        for c in range(1, 18):
            style_cell(ws.cell(row=r, column=c), fill=fill_bg)

    set_col_widths(ws, [2, 22, 10, 8, 8, 10, 8, 8, 10, 8, 8, 10, 8, 8, 2])

    row = 2
    make_title_row(ws, row, 2, "STRENGTH PROGRESSION LOG", 14)
    row += 1
    make_subtitle_row(ws, row, 2, "Track your main lift numbers each week", 14)
    row += 2

    lifts = [
        "Back Squat",
        "Bench Press",
        "Deadlift",
        "OHP / Landmine Press",
        "Weighted Pull-Up",
        "Barbell Row",
        "Front Squat / Goblet Squat",
        "Romanian Deadlift",
    ]

    # For each lift: 12 weeks of Weight x Reps x RPE
    for lift in lifts:
        # Lift name header
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        cell = ws.cell(row=row, column=2, value=lift)
        style_cell(cell, font=Font(name="Calibri", size=12, bold=True, color=ACCENT), fill=fill_raised)
        for c in range(2, 15):
            ws.cell(row=row, column=c).fill = fill_raised
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        # Cycle headers
        for cycle in range(3):
            start_col = 2 + cycle * 4 + (1 if cycle == 0 else 1)

        # Week headers across 12 columns
        cell = ws.cell(row=row, column=2, value="")
        style_cell(cell, font=font_header, fill=fill_header)

        for w in range(12):
            col = 3 + w
            cell = ws.cell(row=row, column=col, value=f"Wk {w+1}")
            style_cell(cell, font=font_header, fill=fill_header)
        row += 1

        # Weight row
        cell = ws.cell(row=row, column=2, value="Weight (kg)")
        style_cell(cell, font=font_body, fill=fill_card, alignment=left_wrap)
        for w in range(12):
            cell = ws.cell(row=row, column=3 + w)
            style_cell(cell, font=font_input, fill=fill_input)
        row += 1

        # Sets x Reps row
        cell = ws.cell(row=row, column=2, value="Sets x Reps")
        style_cell(cell, font=font_body, fill=fill_card, alignment=left_wrap)
        for w in range(12):
            cell = ws.cell(row=row, column=3 + w)
            style_cell(cell, font=font_input, fill=fill_input)
        row += 1

        # RPE row
        cell = ws.cell(row=row, column=2, value="RPE")
        style_cell(cell, font=font_body, fill=fill_card, alignment=left_wrap)
        for w in range(12):
            cell = ws.cell(row=row, column=3 + w)
            style_cell(cell, font=font_input, fill=fill_input)
        row += 1

        # Est. 1RM row (formula)
        cell = ws.cell(row=row, column=2, value="Est. 1RM")
        style_cell(cell, font=font_accent, fill=fill_card, alignment=left_wrap)
        for w in range(12):
            cell = ws.cell(row=row, column=3 + w)
            style_cell(cell, font=font_accent, fill=fill_card)
            # Can't auto-calculate from "sets x reps" text, leave as manual
        row += 2

    # Instructions
    make_subtitle_row(ws, row, 2, "HOW TO USE", 14)
    row += 1
    instructions = [
        "Enter the weight you used for your top/working sets each week",
        "Sets x Reps: e.g., '4x4' or '3x5' or '5x3'",
        "RPE: Rate of Perceived Exertion (6=easy, 7=moderate, 8=hard, 9=very hard, 10=max)",
        "Est. 1RM formula: Weight x (1 + Reps / 30) — calculate for your heaviest set",
    ]
    for inst in instructions:
        cell = ws.cell(row=row, column=2, value=f"  {inst}")
        style_cell(cell, font=font_dim, fill=fill_bg, alignment=left_wrap)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        row += 1

    set_col_widths(ws, [2, 22] + [10] * 12 + [2])
    ws.freeze_panes = "C6"


def build_testing_sheet(wb):
    """Sheet 3: Conditioning test results — baseline + 3 retests."""
    ws = wb.create_sheet("Conditioning Tests")
    ws.sheet_properties.tabColor = RED

    for r in range(1, 40):
        for c in range(1, 14):
            style_cell(ws.cell(row=r, column=c), fill=fill_bg)

    set_col_widths(ws, [2, 6, 28, 14, 14, 14, 14, 14, 14, 14, 30, 2])

    row = 2
    make_title_row(ws, row, 2, "CONDITIONING & FITNESS TESTS", 11)
    row += 1
    make_subtitle_row(ws, row, 2, "Baseline + 3 retests (after each 4-week cycle)", 11)
    row += 2

    # Headers
    headers = [
        "#", "Test", "Target",
        "Baseline\n(Week 0)", "Retest 1\n(Week 4)", "\u0394 1",
        "Retest 2\n(Week 8)", "\u0394 2",
        "Retest 3\n(Week 12)", "\u0394 3", "Notes"
    ]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=2 + i, value=h)
        style_cell(cell, font=font_header, fill=fill_header)
    row += 1

    # Test data
    tests = [
        ("Cardiovascular", [
            ("1", "Resting Heart Rate (bpm)", "< 60 (good)\n< 50 (elite)", "3-day avg"),
            ("2", "2 km Row or Run (time)", "Row < 7:30 / Run < 8:30", "HR finish: ___ / 1-min: ___"),
            ("3", "5-min Assault Bike (cals)", "> 75 (good)\n> 90 (excellent)", "Avg HR: ___"),
        ]),
        ("Strength", [
            ("4", "Back Squat Est. 1RM (kg)", "1.5x BW (good)\n2x BW (excellent)", "3RM weight: ___"),
            ("5", "Bench Press Est. 1RM (kg)", "1x BW (good)\n1.25x BW (excellent)", "3RM weight: ___"),
            ("6", "Deadlift Est. 1RM (kg)", "2x BW (good)\n2.5x BW (excellent)", "3RM weight: ___"),
            ("7", "Max Strict Pull-Ups", "10+ (good)\n15+ (excellent)", ""),
        ]),
        ("Muscular Endurance", [
            ("8", "Max Push-Ups in 2 min", "50+ (good)\n70+ (excellent)", ""),
            ("9", "Max Plank Hold (time)", "2:00+ (good)\n3:00+ (excellent)", "Stop reason: ___"),
        ]),
        ("MMA-Specific", [
            ("10", "3-min Sprawl Test (reps)", "30+ (good)\n40+ (excellent)", ""),
            ("11", "Ground-to-Standing (reps)", "15+ (good)\n20+ (excellent)", ""),
        ]),
        ("Body Composition", [
            ("12", "Bodyweight (kg)", "Track change", "Fasted, morning"),
            ("13", "Waist Circumference (cm)", "Track change", "At navel"),
        ]),
    ]

    for category, items in tests:
        # Category header
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=2, value=category.upper())
        style_cell(cell, font=font_subtitle, fill=fill_raised)
        for c in range(2, 13):
            ws.cell(row=row, column=c).fill = fill_raised
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        for num, name, target, notes in items:
            cell = ws.cell(row=row, column=2, value=num)
            style_cell(cell, font=font_dim, fill=fill_card)

            cell = ws.cell(row=row, column=3, value=name)
            style_cell(cell, font=font_body, fill=fill_card, alignment=left_wrap)

            cell = ws.cell(row=row, column=4, value=target)
            style_cell(cell, font=font_dim, fill=fill_card, alignment=left_wrap)

            # Column layout: E=Baseline, F=Retest1, G=Delta1, H=Retest2, I=Delta2, J=Retest3, K=Delta3
            r = row  # current row number

            # Baseline (input) — col E (5)
            cell = ws.cell(row=r, column=5)
            style_cell(cell, font=font_input, fill=fill_input)

            # Retest 1 (input) — col F (6)
            cell = ws.cell(row=r, column=6)
            style_cell(cell, font=font_input, fill=fill_input)

            # Delta 1 (auto % change: Retest1 vs Baseline) — col G (7)
            cell = ws.cell(row=r, column=7)
            cell.value = f'=IF(AND(E{r}<>"",F{r}<>""),ROUND((F{r}-E{r})/ABS(E{r})*100,1)&"%","")'
            style_cell(cell, font=font_accent, fill=fill_card)
            cell.number_format = '@'

            # Retest 2 (input) — col H (8)
            cell = ws.cell(row=r, column=8)
            style_cell(cell, font=font_input, fill=fill_input)

            # Delta 2 (auto % change: Retest2 vs Retest1) — col I (9)
            cell = ws.cell(row=r, column=9)
            cell.value = f'=IF(AND(F{r}<>"",H{r}<>""),ROUND((H{r}-F{r})/ABS(F{r})*100,1)&"%","")'
            style_cell(cell, font=font_accent, fill=fill_card)
            cell.number_format = '@'

            # Retest 3 (input) — col J (10)
            cell = ws.cell(row=r, column=10)
            style_cell(cell, font=font_input, fill=fill_input)

            # Delta 3 (auto % change: Retest3 vs Retest2) — col K (11)
            cell = ws.cell(row=r, column=11)
            cell.value = f'=IF(AND(H{r}<>"",J{r}<>""),ROUND((J{r}-H{r})/ABS(H{r})*100,1)&"%","")'
            style_cell(cell, font=font_accent, fill=fill_card)
            cell.number_format = '@'

            # Notes
            cell = ws.cell(row=row, column=12, value=notes)
            style_cell(cell, font=font_dim, fill=fill_card, alignment=left_wrap)

            row += 1

    row += 1
    # 1RM formula reference
    make_subtitle_row(ws, row, 2, "1RM CALCULATION: Estimated 1RM = Weight x (1 + Reps / 30)", 11)
    row += 1
    cell = ws.cell(row=row, column=2, value="  Example: 3RM of 100 kg = 100 x (1 + 3/30) = 110 kg.  Always use a 3RM for consistency.")
    style_cell(cell, font=font_dim, fill=fill_bg, alignment=left_wrap)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)

    ws.freeze_panes = "D7"


def build_body_tracker(wb):
    """Sheet 4: Weekly body composition and recovery tracker."""
    ws = wb.create_sheet("Weekly Check-In")
    ws.sheet_properties.tabColor = GREEN

    for r in range(1, 30):
        for c in range(1, 18):
            style_cell(ws.cell(row=r, column=c), fill=fill_bg)

    set_col_widths(ws, [2, 20] + [10] * 12 + [2])

    row = 2
    make_title_row(ws, row, 2, "WEEKLY CHECK-IN", 14)
    row += 1
    make_subtitle_row(ws, row, 2, "Fill in every Sunday — track trends, not single data points", 14)
    row += 2

    # Week headers
    cell = ws.cell(row=row, column=2, value="")
    style_cell(cell, font=font_header, fill=fill_header)
    for w in range(12):
        cell = ws.cell(row=row, column=3 + w, value=f"Week {w+1}")
        style_cell(cell, font=font_header, fill=fill_header)
    row += 1

    metrics = [
        ("Body Composition", [
            ("Bodyweight (kg)", "Fasted, morning, same scale"),
            ("Waist (cm)", "At navel, relaxed"),
        ]),
        ("Recovery & Wellness", [
            ("Avg Sleep (hrs)", "Target: 8-9 hours"),
            ("Sleep Quality (1-10)", "1=terrible, 10=perfect"),
            ("Energy Level (1-10)", "Average for the week"),
            ("Muscle Soreness (1-10)", "1=none, 10=crippling"),
            ("Motivation (1-10)", "How driven were you?"),
            ("Stress Level (1-10)", "Life + training combined"),
        ]),
        ("Training Summary", [
            ("Sessions Completed", "Out of 12 planned"),
            ("Sessions Missed", "Log reason in notes"),
            ("Hardest Session", "Which one pushed you most?"),
        ]),
        ("Nutrition Compliance", [
            ("Protein Target Hit? (Y/N)", "Days you hit target / 7"),
            ("Hydration (L/day avg)", "Target: 3-4L"),
            ("Supplements Taken? (Y/N)", "Creatine, omega-3, etc."),
        ]),
    ]

    for category, items in metrics:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        cell = ws.cell(row=row, column=2, value=category.upper())
        style_cell(cell, font=font_subtitle, fill=fill_raised)
        for c in range(2, 15):
            ws.cell(row=row, column=c).fill = fill_raised
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        for name, hint in items:
            cell = ws.cell(row=row, column=2, value=name)
            style_cell(cell, font=font_body, fill=fill_card, alignment=left_wrap)

            for w in range(12):
                cell = ws.cell(row=row, column=3 + w)
                style_cell(cell, font=font_input, fill=fill_input)
            row += 1

    row += 1
    # Notes section
    make_subtitle_row(ws, row, 2, "WEEKLY NOTES / ADJUSTMENTS", 14)
    row += 1
    cell = ws.cell(row=row, column=2, value="")
    style_cell(cell, font=font_header, fill=fill_header)
    for w in range(12):
        cell = ws.cell(row=row, column=3 + w, value=f"Week {w+1}")
        style_cell(cell, font=font_header, fill=fill_header)
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row + 3, end_column=2)
    cell = ws.cell(row=row, column=2, value="Notes")
    style_cell(cell, font=font_body, fill=fill_card, alignment=left_wrap)
    for r_off in range(4):
        for w in range(12):
            cell = ws.cell(row=row + r_off, column=3 + w)
            style_cell(cell, font=font_input, fill=fill_input, alignment=left_top)

    ws.freeze_panes = "C6"


def main():
    wb = openpyxl.Workbook()

    build_daily_tracker(wb)
    build_strength_log(wb)
    build_testing_sheet(wb)
    build_body_tracker(wb)

    import os
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "MMA_Training_Tracker.xlsx")
    wb.save(out)
    size = os.path.getsize(out)
    print(f"Built MMA_Training_Tracker.xlsx: {size:,} bytes ({size/1024:.0f} KB)")


if __name__ == "__main__":
    main()
